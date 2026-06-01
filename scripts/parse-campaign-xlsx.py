#!/usr/bin/env python3
"""
parse_campaign_xlsx.py — unified parser for bar crawl campaign xlsx files.

Handles three formats:
  1. CLUSTER (Halloween 2025, NYE 2026)
       Sheets named "Calgary, AB" etc. with "CLUSTER 1/2/3" headers,
       each followed by a Venue Type | Venue Name | Email | Contact
       Name | Contact # | Hours | Address | Capacity | Specials |
       Notes | Confirmation row.

  2. MULTI-CRAWL (SPD 2026)
       Same per-cluster layout but with "Friday Crawl 1/2",
       "Saturday Crawl 1/2/3" headers (variable count). Adds an
       "Internal Contact" column we skip (operator-only).

  3. LEGACY (Halloween 2024, NYE 2025, SPD 2025)
       Single-cluster (or multi-day single-cluster) with "Wristband
       Venue / Alt Venue N / Final Venue" or "Venue Slot N" labels.
       Different column order (Phone before Email). Some have
       multiple day sections within one sheet.

All formats emit the same JSON shape so the import orchestrator
can iterate uniformly. Per-city structure:

  {
    "city_label": "Calgary, AB",
    "confirmed_venues": [
      {
        cluster_num, date_label, slot_role, slot_position,
        venue_type_raw, venue_name, venue_email, contact_name,
        contact_phone, proposed_hours, address, capacity,
        specials, notes, confirmation
      },
      ...
    ],
    "warm_leads":   [...],   # status_note + standard venue fields
    "cold_outreach": [...]   # venue_name + status_raw + email +
                             # phone + other_contact + hours +
                             # alt_email + notes
  }

Run:
  python3 parse_campaign_xlsx.py <input.xlsx> <output.json> [format]

`format` is auto-detected if omitted; pass "cluster" / "multi" /
"legacy" to force.
"""

import json
import re
import sys
from openpyxl import load_workbook

# ----------------------------------------------------------------
# Sheets we ALWAYS skip — operator-internal trackers / tools
# ----------------------------------------------------------------
SKIP_SHEET_NAMES = {
    "tracker", "trackers", "to cancel", "wb tracker", "hosts",
    "cancellation cities", "spd sales data", "previous spd",
    "cmnd sheet", "event submission", "do not use", "email templates",
    "host check", "city tracking", "priority", "form responses 1",
    "manual host compensation", "wristband trkr", "contact",
    "nye sales data", "code", "prio data", "raw confirmed",
    "hosts wages", "toronto pricing", "sheet18",
    "toronto venue hitlist", "competitors", "toronto contacts",
    "intl crawl tracker", "eventbrite checklist", "church st. bars",
    "outreach emails", "toronto crawls spd", "toronto crawl tracker",
    "toronto nye tracker", "template",
    # Toronto-format sheets — grid layout, parsed separately
    "toronto", "toronto crawls", "toronto vday",
}

def is_skippable(sheet_name: str) -> bool:
    lower = sheet_name.lower().strip()
    if lower in SKIP_SHEET_NAMES:
        return True
    # Heuristic — sheets that look like internal trackers
    if "tracker" in lower or "checklist" in lower or "template" in lower:
        return True
    return False


# ----------------------------------------------------------------
# Cell utilities
# ----------------------------------------------------------------
def cell_str(v) -> str:
    """Normalize a cell value to a stripped string. None/empty → ''."""
    if v is None:
        return ""
    s = str(v).strip()
    # Drop the trailing ".0" Excel adds to int-like floats
    if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
        s = s[:-2]
    return s


def cell_clean_phone(v) -> str:
    """Phone normalization — strip Excel's ".0" trailing on numerics."""
    s = cell_str(v)
    if not s:
        return ""
    # If it's all digits with a trailing .0, strip it (Excel float coercion)
    return s


def cell_int(v):
    """Try to interpret a cell as an int (for capacity). Returns None if not."""
    if v is None:
        return None
    try:
        f = float(v)
        if f != f:  # NaN
            return None
        if f <= 0:
            return None
        return int(f)
    except (ValueError, TypeError):
        return None


# ----------------------------------------------------------------
# Format detection
# ----------------------------------------------------------------
def detect_format(ws) -> str:
    """
    Scan the first ~15 rows + first 6 columns. Return one of:
      'cluster'  — CLUSTER 1/2/3 headers
      'multi'    — Friday/Saturday Crawl N headers
      'legacy'   — Wristband Venue / Alt Venue / Final Venue / Venue Slot
      'unknown'  — nothing matched
    """
    found_cluster = False
    found_crawl = False
    found_legacy = False

    for r in range(1, 18):
        for c in range(1, 8):
            v = cell_str(ws.cell(r, c).value).upper()
            if not v:
                continue
            if re.match(r"^CLUSTER\s+\d+$", v):
                found_cluster = True
            if re.match(r"^(FRIDAY|SATURDAY|SUNDAY|THURSDAY)\s+CRAWL\s*\d+", v):
                found_crawl = True
            if re.match(r"^(WRISTBAND\s*VENUE|ALT\s*VENUE\s*\d+|FINAL\s*VENUE|VENUE\s*SLOT\s*\d+)", v):
                found_legacy = True

    if found_crawl:
        return "multi"
    if found_cluster:
        return "cluster"
    if found_legacy:
        return "legacy"
    return "unknown"


# ----------------------------------------------------------------
# Column-position mapping from a header row
# ----------------------------------------------------------------
HEADER_ALIASES = {
    "venue_name": ["venue name", "scheduled venue name", "venues", "venue"],
    "venue_email": ["venue email", "contact email", "email"],
    "contact_name": ["contact name", "internal contact"],
    "contact_phone": ["contact #", "contact phone", "phone", "phone #",
                      "contact phone number"],
    "proposed_hours": ["proposed hours", "proposed venue hours", "hours"],
    "address": ["address", "venue address"],
    "capacity": ["capacity", "venue capacity"],
    "specials": ["specials", "venue drink specials", "drink specials"],
    "notes": ["notes", "venue notes"],
    "confirmation": ["confirmation", "venue confirmed",
                     "venue confirmed w/written agreement", "confirmed?",
                     "confirmed"],
}

def map_columns(ws, header_row: int) -> dict:
    """
    Given a header row index, return a dict of field-name → column-index
    (1-based) by matching header text against HEADER_ALIASES.

    Headers can wrap or have trailing whitespace; we lowercase + normalize
    whitespace before matching.
    """
    mapping = {}
    # Scan up to 16 columns for headers
    for c in range(1, 16):
        raw = ws.cell(header_row, c).value
        if raw is None:
            continue
        normalized = re.sub(r"\s+", " ", str(raw).lower().strip())
        # Strip trailing /.. fragments that operators add
        normalized = normalized.split("\n")[0].strip()
        for field, aliases in HEADER_ALIASES.items():
            if normalized in aliases and field not in mapping:
                mapping[field] = c
                break
    return mapping


# ----------------------------------------------------------------
# Slot-role normalization
# ----------------------------------------------------------------
def normalize_slot(venue_type_raw: str) -> tuple:
    """
    Convert a raw venue-type label to a (slot_role, slot_position) tuple.
    Returns (None, None) for unknown types (we still capture the row but
    don't tag it for the events/venue_events table).

    Roles map to the venueRole pg enum:
      wristband, middle, final, alt_final
    """
    if not venue_type_raw:
        return (None, None)
    s = venue_type_raw.upper().strip()

    if s == "WRISTBAND" or "WRISTBAND" in s:
        return ("wristband", 1)
    if s == "FINAL" or s == "FINAL VENUE":
        return ("final", 1)
    m = re.match(r"^ALT\s*FINAL\s*(\d+)?$", s)
    if m:
        return ("alt_final", int(m.group(1) or "1"))
    m = re.match(r"^PARTICIPATING\s*(\d+)$", s)
    if m:
        return ("middle", int(m.group(1)))
    m = re.match(r"^OVERFLOW\s*FINAL$", s)
    if m:
        return ("alt_final", 1)
    # LEGACY mappings
    m = re.match(r"^ALT\s*VENUE\s*(\d+)$", s)
    if m:
        return ("middle", int(m.group(1)))
    m = re.match(r"^VENUE\s*SLOT\s*(\d+)$", s)
    if m:
        return ("middle", int(m.group(1)))
    return (None, None)


# ----------------------------------------------------------------
# Block scanning utilities
# ----------------------------------------------------------------
def find_section_starts(ws, max_row: int = 200) -> list:
    """
    Walk the sheet and find the START rows of each section.
    Returns a list of (section_type, header_label, header_row, data_start_row).

    section_type is one of:
      'cluster'     — confirmed venues block
      'warm'        — warm leads block
      'cold'        — cold outreach block

    The 'header_row' is the row with the column titles ("Venue Type",
    "Venue Name", etc.). data_start_row is header_row + 1.
    """
    sections = []
    # We scan row by row looking for keyword markers in column A,
    # then look ahead for the actual column-header row.
    last_row = min(ws.max_row or max_row, max_row)

    r = 1
    while r <= last_row:
        a = cell_str(ws.cell(r, 1).value).upper()
        b = cell_str(ws.cell(r, 2).value).lower()

        # CLUSTER markers
        if re.match(r"^CLUSTER\s+\d+", a):
            header_row = _find_header_row(ws, r + 1, max_look=4)
            if header_row:
                sections.append(("cluster", a, header_row, header_row + 1))
        # MULTI-CRAWL markers: "Friday Crawl 1", etc.
        elif re.match(r"^(FRIDAY|SATURDAY|SUNDAY|THURSDAY)\s+CRAWL", a):
            header_row = _find_header_row(ws, r + 1, max_look=4)
            if header_row:
                sections.append(("cluster", a, header_row, header_row + 1))
        # LEGACY day markers (single cluster per day in older sheets)
        elif re.match(r"^(SATURDAY|FRIDAY|SUNDAY)[ ,]+(MARCH|APRIL|MAY|JUNE|OCTOBER|NOVEMBER|DECEMBER|JANUARY|FEBRUARY)", a):
            header_row = _find_header_row(ws, r + 1, max_look=4)
            if header_row:
                sections.append(("cluster", a, header_row, header_row + 1))
        # WARM LEADS marker
        elif a in ("WARM LEADS", "WARM LEAD") or "warm lead" in a.lower():
            header_row = _find_header_row(ws, r + 1, max_look=4)
            if header_row:
                sections.append(("warm", a, header_row, header_row + 1))
        # COLD OUTREACH marker
        elif "COLD OUTREACH" in a or a == "COLD":
            header_row = _find_header_row(ws, r + 1, max_look=4)
            if header_row:
                sections.append(("cold", a, header_row, header_row + 1))
        r += 1

    # FALLBACK 1: sheets with NO explicit cluster header at all (e.g.
    # some legacy sheets jump straight to "Wristband Venue" in row 3).
    # Treat the whole block as a single cluster labeled "Day 1".
    if not any(s[0] == "cluster" for s in sections):
        # Scan row 2-6 for a header with venue name in column B
        for hr in range(2, 8):
            if _looks_like_header(ws, hr):
                sections.insert(0, ("cluster", "Day 1", hr, hr + 1))
                break

    # FALLBACK 2: sheets where the FIRST section has no day header
    # but a LATER section does (e.g. SPD 2025 sheets that start with
    # an un-labeled Friday block followed by a "Saturday March 15th"
    # header for the second day). Without this, the Friday block is
    # lost — only Saturday gets captured. Detection: there's an
    # early header row (in the first 8 rows) that precedes the
    # FIRST cluster section start.
    cluster_sections = [s for s in sections if s[0] == "cluster"]
    if cluster_sections:
        first_cluster_row = cluster_sections[0][2]  # header_row of first cluster
        if first_cluster_row > 8:
            # Look for a header row in rows 2-7
            for hr in range(2, 8):
                if _looks_like_header(ws, hr) and hr < first_cluster_row:
                    sections.insert(0, ("cluster", "Day 1", hr, hr + 1))
                    break

    return sections


def _looks_like_header(ws, row: int) -> bool:
    """
    A row counts as a header if it contains "Venue Name" or similar in
    column B (or some adjacent column).
    """
    for c in range(1, 8):
        v = cell_str(ws.cell(row, c).value).lower()
        if v in ("venue name", "scheduled venue name", "venues", "venue"):
            return True
    return False


def _find_header_row(ws, start_row: int, max_look: int = 4):
    """
    Starting from start_row, look for a 'Venue Name' header within the
    next max_look rows. Returns the row number or None.
    """
    for r in range(start_row, start_row + max_look):
        if _looks_like_header(ws, r):
            return r
    return None


# ----------------------------------------------------------------
# Section extraction
# ----------------------------------------------------------------
def extract_cluster_rows(ws, header_row: int, data_start: int,
                         cluster_num: int, date_label: str,
                         next_section_row: int) -> list:
    """
    Extract the venue rows in one cluster section.
    Reads rows from data_start until next_section_row (exclusive) OR
    until we hit 3 consecutive blank rows.
    """
    cols = map_columns(ws, header_row)
    if "venue_name" not in cols:
        return []  # malformed section

    rows = []
    blank_streak = 0
    r = data_start
    while r < next_section_row and r <= (ws.max_row or 200):
        venue_name = cell_str(ws.cell(r, cols["venue_name"]).value)
        venue_type_raw = cell_str(ws.cell(r, 1).value)
        if not venue_name and not venue_type_raw:
            blank_streak += 1
            if blank_streak >= 3:
                break
            r += 1
            continue
        if not venue_name:
            # Type-only row (a placeholder slot) — skip
            r += 1
            blank_streak = 0
            continue
        blank_streak = 0

        slot_role, slot_position = normalize_slot(venue_type_raw)
        rows.append({
            "cluster_num": cluster_num,
            "date_label": date_label,
            "slot_role": slot_role,
            "slot_position": slot_position,
            "venue_type_raw": venue_type_raw,
            "venue_name": venue_name,
            "venue_email": cell_str(ws.cell(r, cols["venue_email"]).value) if "venue_email" in cols else None,
            "contact_name": cell_str(ws.cell(r, cols["contact_name"]).value) if "contact_name" in cols else None,
            "contact_phone": cell_clean_phone(ws.cell(r, cols["contact_phone"]).value) if "contact_phone" in cols else None,
            "proposed_hours": cell_str(ws.cell(r, cols["proposed_hours"]).value) if "proposed_hours" in cols else None,
            "address": cell_str(ws.cell(r, cols["address"]).value) if "address" in cols else None,
            "capacity": cell_int(ws.cell(r, cols["capacity"]).value) if "capacity" in cols else None,
            "specials": cell_str(ws.cell(r, cols["specials"]).value) if "specials" in cols else None,
            "notes": cell_str(ws.cell(r, cols["notes"]).value) if "notes" in cols else None,
            "confirmation": cell_str(ws.cell(r, cols["confirmation"]).value) if "confirmation" in cols else None,
        })
        r += 1
    # Strip empty-string fields to None for cleaner JSON
    return [_clean_nones(d) for d in rows]


def extract_warm_rows(ws, header_row: int, data_start: int,
                      next_section_row: int) -> list:
    """Extract warm-leads section rows."""
    cols = map_columns(ws, header_row)
    if "venue_name" not in cols:
        return []

    rows = []
    blank_streak = 0
    r = data_start
    while r < next_section_row and r <= (ws.max_row or 200):
        venue_name = cell_str(ws.cell(r, cols["venue_name"]).value)
        status_note = cell_str(ws.cell(r, 1).value)
        if not venue_name:
            blank_streak += 1
            if blank_streak >= 3:
                break
            r += 1
            continue
        blank_streak = 0
        rows.append({
            "status_note": status_note or None,
            "venue_name": venue_name,
            "venue_email": cell_str(ws.cell(r, cols["venue_email"]).value) if "venue_email" in cols else None,
            "contact_name": cell_str(ws.cell(r, cols["contact_name"]).value) if "contact_name" in cols else None,
            "contact_phone": cell_clean_phone(ws.cell(r, cols["contact_phone"]).value) if "contact_phone" in cols else None,
            "proposed_hours": cell_str(ws.cell(r, cols["proposed_hours"]).value) if "proposed_hours" in cols else None,
            "address": cell_str(ws.cell(r, cols["address"]).value) if "address" in cols else None,
            "capacity": cell_int(ws.cell(r, cols["capacity"]).value) if "capacity" in cols else None,
            "specials": cell_str(ws.cell(r, cols["specials"]).value) if "specials" in cols else None,
            "notes": cell_str(ws.cell(r, cols["notes"]).value) if "notes" in cols else None,
            "confirmation": cell_str(ws.cell(r, cols["confirmation"]).value) if "confirmation" in cols else None,
        })
        r += 1
    return [_clean_nones(d) for d in rows]


def extract_cold_rows(ws, header_row: int, data_start: int) -> list:
    """
    Cold outreach section runs to end of sheet. We stop on 5 consecutive
    blank rows.
    """
    cols = map_columns(ws, header_row)
    if "venue_name" not in cols:
        return []

    rows = []
    blank_streak = 0
    last_row = min(ws.max_row or 300, 500)
    r = data_start
    while r <= last_row:
        venue_name = cell_str(ws.cell(r, cols["venue_name"]).value)
        if not venue_name:
            blank_streak += 1
            if blank_streak >= 5:
                break
            r += 1
            continue
        blank_streak = 0
        rows.append({
            "venue_name": venue_name,
            "status_raw": cell_str(ws.cell(r, 1).value) or None,  # col A
            "venue_email": cell_str(ws.cell(r, cols["venue_email"]).value) if "venue_email" in cols else None,
            "contact_name": cell_str(ws.cell(r, cols["contact_name"]).value) if "contact_name" in cols else None,
            "phone": cell_clean_phone(ws.cell(r, cols["contact_phone"]).value) if "contact_phone" in cols else None,
            "address": cell_str(ws.cell(r, cols["address"]).value) if "address" in cols else None,
            "notes": cell_str(ws.cell(r, cols["notes"]).value) if "notes" in cols else None,
        })
        r += 1
    return [_clean_nones(d) for d in rows]


def _clean_nones(d: dict) -> dict:
    """Convert empty strings to None in a dict (cleaner JSON)."""
    return {k: (v if (v not in ("", None) and v == v) else None)
            for k, v in d.items()}


# ----------------------------------------------------------------
# Per-sheet parse
# ----------------------------------------------------------------
def parse_sheet(ws, sheet_name: str) -> dict:
    """
    Parse a single city sheet into the unified output shape.
    Returns None if the sheet has no recognizable structure.
    """
    sections = find_section_starts(ws)
    if not sections:
        return None

    # Group sections by type
    cluster_sections = [s for s in sections if s[0] == "cluster"]
    warm_sections = [s for s in sections if s[0] == "warm"]
    cold_sections = [s for s in sections if s[0] == "cold"]

    # Sort all sections by row so we can compute "next_section_row"
    sections.sort(key=lambda s: s[2])

    confirmed_venues = []
    for idx, (_kind, label, hr, ds) in enumerate(
        [s for s in sections if s[0] == "cluster"]
    ):
        # Find the next section START to bound this one
        next_row = (ws.max_row or 300) + 1
        for s2 in sections:
            if s2[2] > hr:
                next_row = s2[2]
                break
        rows = extract_cluster_rows(ws, hr, ds, idx + 1, label, next_row)
        confirmed_venues.extend(rows)

    warm_leads = []
    for (_kind, label, hr, ds) in warm_sections:
        # Bound at next section start
        next_row = (ws.max_row or 300) + 1
        for s2 in sections:
            if s2[2] > hr:
                next_row = s2[2]
                break
        warm_leads.extend(extract_warm_rows(ws, hr, ds, next_row))

    cold_outreach = []
    for (_kind, label, hr, ds) in cold_sections:
        cold_outreach.extend(extract_cold_rows(ws, hr, ds))

    return {
        "city_label": sheet_name,
        "confirmed_venues": confirmed_venues,
        "warm_leads": warm_leads,
        "cold_outreach": cold_outreach,
    }


# ----------------------------------------------------------------
# Main
# ----------------------------------------------------------------
def parse_workbook(path: str) -> dict:
    # Non-readonly is slower to load but much faster per-cell — and these
    # xlsx files are all under 6MB so the upfront cost is fine.
    wb = load_workbook(path, data_only=True, read_only=False)
    output = {}
    summary = []
    for sn in wb.sheetnames:
        if is_skippable(sn):
            summary.append((sn, "SKIPPED"))
            continue
        try:
            ws = wb[sn]
            fmt = detect_format(ws)
            if fmt == "unknown":
                summary.append((sn, "UNKNOWN format"))
                continue
            parsed = parse_sheet(ws, sn)
            if not parsed:
                summary.append((sn, f"{fmt} — NO sections"))
                continue
            cv = len(parsed["confirmed_venues"])
            wl = len(parsed["warm_leads"])
            co = len(parsed["cold_outreach"])
            summary.append((sn, f"{fmt} — {cv} conf / {wl} warm / {co} cold"))
            output[sn] = parsed
        except Exception as e:
            summary.append((sn, f"FAILED: {e}"))
    wb.close()
    return output, summary


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: parse_campaign_xlsx.py <input.xlsx> <output.json>")
        sys.exit(1)
    in_path = sys.argv[1]
    out_path = sys.argv[2]
    output, summary = parse_workbook(in_path)
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)
    print(f"\nParsed {in_path}")
    print(f"  Sheets processed: {len(summary)}")
    print(f"  Cities extracted: {len(output)}")
    print(f"\nPer-sheet summary:")
    for sn, status in summary:
        print(f"  {sn:35s}  {status}")
    print(f"\nOutput → {out_path}")
